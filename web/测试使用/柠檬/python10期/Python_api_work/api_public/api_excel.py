from openpyxl import load_workbook
from Python_api_work.api_public.api_config import Config
from Python_api_work.api_public import api_path
#生成一个读取和写入Excel表格的类
class Do_Excel:
    def __init__(self,file):
        self.file=file
        self.sheet_list=eval(Config().config(api_path.config_path,'CASECONFIG','sheet_list'))
    #读取
    def read(self):

        work_book=load_workbook(self.file)
        test_data = []
        for sheet_name in self.sheet_list:
            sheet=work_book[sheet_name]
            #读取配置文件中执行的用例
            if self.sheet_list[sheet_name]=='all':
                #读取标题
                key=[]
                for item in range(1,sheet.max_column-1):
                    key.append(sheet.cell(1,item).value)
                #读取数据
                for row in range(2,sheet.max_row+1):
                    sub_data={}
                    sub_data['sheet']=sheet_name
                    for column in range(1,sheet.max_column-1):
                        sub_data[key[column-1]]=sheet.cell(row,column).value
                    test_data.append(sub_data)
            else:
                #读取标题
                key = []
                for item in range(1, sheet.max_column - 1):
                    key.append(sheet.cell(1, item).value)
                #读取数据
                for row in self.sheet_list[sheet_name]:
                    sub_data={}
                    sub_data['sheet']=sheet_name
                    for column in range(1,sheet.max_column-1):
                        sub_data[key[column-1]]=sheet.cell(row+1,column).value
                    test_data.append(sub_data)
        return test_data

    #写回函数
    def write(self,sheet,row,actual,result):
        work_book=load_workbook(self.file)
        sheet=work_book[sheet]
        sheet.cell(row,sheet.max_column-1).value=actual
        sheet.cell(row,sheet.max_column).value=result
        work_book.save(self.file)

