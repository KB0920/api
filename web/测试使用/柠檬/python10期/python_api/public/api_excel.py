from openpyxl import load_workbook
from python_api.public.api_config import Config
from python_api.public import api_path

class Doexcel:
    def __init__(self,file):
        self.file=file
        self.sheet_list = eval(Config().config(api_path.config_path, 'CASECONFIG', 'sheet_list'))

    def read(self):
        test_data = []
        wb = load_workbook(self.file)
        for h in self.sheet_list:
            st=wb[h]
            if self.sheet_list[h]=='all':
                key=[]
                for item in range(1,st.max_column-1):
                    key.append(st.cell(1,item).value)

                for i in range(2,st.max_row+1):
                    sub_data={}
                    sub_data['sheet_name'] = h
                    for j in range(1,st.max_column-1):
                        sub_data[key[j-1]]=st.cell(i,j).value
                    test_data.append(sub_data)
            else:
                key = []
                for item in range(1, st.max_column - 1):
                    key.append(st.cell(1, item).value)

                for i in self.sheet_list[h]:
                    sub_data = {}
                    sub_data['sheet_name']=h
                    for j in range(1, st.max_column - 1):
                        sub_data[key[j - 1]] = st.cell(i+1, j).value

                    test_data.append(sub_data)

        return test_data
        # if mode=='on':
        #     finally_data=test_data
        # else:
        #     finally_data=[]
        #     for item in test_data:
        #         if item['id'] in case_id:
        #             finally_data.append(item)
        # return finally_data
    def write(self,sheet_name,row,actual,result):
        wb = load_workbook(self.file)
        st=wb[sheet_name]

        st.cell(row,st.max_column-1).value=actual
        st.cell(row,st.max_column).value=result
        wb.save(self.file)


if __name__ == '__main__':
    # sheet_list = eval(Config().config(api_path.config_path, 'CASECONFIG', 'sheet_list'))
    # print(sheet_list)
    # for sheet in sheet_list:
    #     print(sheet)
    res=Doexcel(api_path.data_path).read()
    # for key in sheet_list:
    #     print(sheet_list[key])
    for i in res:
        print(i)
