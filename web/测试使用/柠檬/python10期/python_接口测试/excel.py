from openpyxl import load_workbook
#第一阶段
#1、完成单一接口的请求
#1) request 单一请求完成，封装成类
#2) 要加入异常处理
#3) 测试数据的批量写入
#sheet.max_row sheet.max_column

#2、测试数据和测试结果的批量操作和批量写回
#1) 可以根据code，msg进行判断
#'http://119.23.241.154:8080/futureloan/mvc/api/member/register'
class DoExcel:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet
    def read(self):

        wb=load_workbook(self.file)
        st=wb[self.sheet]

        key=[]
        for i in range(1,st.max_row-1):
            key.append(st.cell(1,i).value)

        data=[]
        for i in range(2,st.max_row+1):
            dict={}
            for j in range(1,7):
                dict[key[j-1]]=st.cell(i,j).value
            data.append(dict)
        return data

    def write(self,row,actual,result):
        wb=load_workbook(self.file)
        st=wb[self.sheet]

        st.cell(row,7).value=actual
        st.cell(row,8).value=result
        wb.save(self.file)

if __name__ == '__main__':
    DoExcel().read('API.xlsx','API')


