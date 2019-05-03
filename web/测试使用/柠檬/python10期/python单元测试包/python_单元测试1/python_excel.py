from openpyxl import load_workbook
class DoExcel:
    def __init__(self,file,sheet):
        self.file=file
        self.sheet=sheet

    def read(self,mode,case_id):
        wb=load_workbook(self.file)
        st=wb[self.sheet]

        key=[]
        for i in range(1,6):
            key.append(st.cell(1,i).value)

        test_data=[]
        for i in range(2,8):
            test={}
            for j in range(1,6):
                test[key[j-1]]=st.cell(i,j).value
            test_data.append(test)

        if mode=='on':
            finally_data=test_data
        else:
            finally_data=[]
            for item in test_data:
                if item['id'] in case_id:
                    finally_data.append(item)

        return finally_data

    def write(self,row,actual,result):
        wb = load_workbook(self.file)
        st = wb[self.sheet]

        st.cell(row,6).value=actual
        st.cell(row,7).value=result
        wb.save(self.file)