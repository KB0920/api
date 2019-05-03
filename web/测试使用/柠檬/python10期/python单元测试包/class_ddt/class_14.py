__author__ = 'Administrator'
#pip install openpyxl
#文件必须是xlsx结尾的才可以读取


from openpyxl import load_workbook
class DoExcel:
    def __init__(self,xlsx,sheet):
        self.xlsx=xlsx
        self.sheet=sheet

    def read(self,mode,case_id_list):
        wb=load_workbook(self.xlsx)
        st=wb[self.sheet]

        testdata=[]
        header=[]
        for i in range(1,6):
            header.append(st.cell(1,i).value)
        for i in range(2,8):
            list={}
            for j in range(1,6):
                list[header[j-1]]=st.cell(i,j).value
            testdata.append(list)

        if mode=='on':#如果是on就执行所有用例
            finally_data=testdata
        else:
            finally_data=[]
            for item in testdata:#存储最终要执行的用例
                if item['id'] in case_id_list:
                    finally_data.append(item)
        return finally_data




    def write(self,row,actual,result):
        wb=load_workbook(self.xlsx)
        st=wb[self.sheet]
        #写入结果
        st.cell(row,6).value=actual
        st.cell(row,7).value=result
        wb.save(self.xlsx)







# if __name__ == '__main__':
    # testdata=DoExcel().read()

