# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from webservice.Common.config import Config
from webservice.Common import path

class Do_Excel:

    def __init__(self,file):
        self.file=file
        self.config=Config(path.conf_path)

    def write(self,sheetname,row,actual,result):
        workbook=load_workbook(self.file)
        sheet=workbook[sheetname]

        sheet.cell(row,sheet.max_column-1).value=actual
        sheet.cell(row,sheet.max_column).value=result

        workbook.save(self.file)

    def read(self,sheetname):
        workbook=load_workbook(self.file)
        sheet = workbook[sheetname]
        test_data=[]
        if self.config.config_get('CASE',sheetname)=='all':
            key=[]
            for column in range(1,sheet.max_column-1):
                key.append(sheet.cell(1,column).value)

            for row in range(2,sheet.max_row+1):
                sub_data={}
                sub_data['sheet_name'] = sheetname
                for column in range(1,sheet.max_column-1):
                    sub_data[key[column-1]]=sheet.cell(row,column).value
                test_data.append(sub_data)
        else:
            key=[]
            for column in range(1,sheet.max_column-1):
                key.append(sheet.cell(1,column).value)

            for row in eval(self.config.config_get('CASE',sheetname)):
                sub_data={}
                sub_data['sheet_name'] = sheetname
                for column in range(1,sheet.max_column-1):
                    sub_data[key[column-1]]=sheet.cell(row+1,column).value
                test_data.append(sub_data)
        return test_data


