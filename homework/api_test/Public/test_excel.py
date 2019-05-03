# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from api_test.Public import test_path
from api_test.Public.test_config import Config

class Do_Excel:

    def __init__(self,file):
        self.file=file
        self.config=Config(test_path.config_path)

    def write(self,sheet_name,row,actual,result):
        workbook=load_workbook(self.file)
        sheet=workbook[sheet_name]

        sheet.cell(row,sheet.max_column-1).value=actual
        sheet.cell(row,sheet.max_column).value=result
        workbook.save(self.file)

    def read(self,sheet_name):
        workbook=load_workbook(self.file)
        sheet=workbook[sheet_name]
        test_data = []
        if self.config.config_get('CASE',sheet_name)=='all':
            key = []
            for column in range(1,sheet.max_column-1):
                key.append(sheet.cell(1,column).value)

            for row in range(2,sheet.max_row+1):
                sub_data = {}
                sub_data['sheet_name']=sheet_name
                for column in range(1,sheet.max_column-1):
                    sub_data[key[column-1]]=sheet.cell(row,column).value
                test_data.append(sub_data)

        else:
            key=[]
            for column in range(1,sheet.max_column-1):
                key.append(sheet.cell(1,column).value)

            for row in eval(self.config.config_get('CASE',sheet_name)):
                sub_data={}
                sub_data['sheet_name']=sheet_name
                for column in range(1,sheet.max_column-1):
                    sub_data[key[column-1]]=sheet.cell(row+1,column).value
                test_data.append(sub_data)

        return test_data



if __name__ == '__main__':
    # result=Config(test_path.config_path).config_get('CASE','login')
    # print(result)
    result=Do_Excel(test_path.excel_path).read('register')
    for i in result:
        print(i)